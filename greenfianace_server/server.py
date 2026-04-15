from __future__ import annotations

<<<<<<< HEAD
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
=======
import os
from pathlib import Path
from typing import Any, Optional
from functools import lru_cache

from dotenv import load_dotenv
from dbutils.pooled_db import PooledDB
from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import load_workbook
import pymysql
>>>>>>> origin/main

from ai_agent import stream_chat_agent_events, stream_summary_agent_events
from ai_context import build_chat_messages, build_summary_messages
from ai_service import AiServiceError, request_deepseek_completion
from ai_types import AiChatRequest, AiSummaryRequest
from data_service import (
    get_city_rows,
    get_macro_rows,
    get_macro_stats_records,
    get_predict_payload,
    get_province_rows,
)

app = FastAPI(title="绿色金融碳减排可视化大屏 API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _county_carbon_xlsx_path() -> Path:
    candidates: list[Path] = []
    env_path = os.getenv("COUNTY_CARBON_XLSX")
    if env_path:
        candidates.append(Path(env_path))
    candidates.extend(
        [
            ROOT / "data" / "2000-2024县级碳排放(1).xlsx",
            Path("D:/CountyCarbonEmissions/2000-2024县级碳排放(1).xlsx"),
        ]
    )
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


def _city_carbon_xlsx_path() -> Path:
    candidates: list[Path] = []
    env_path = os.getenv("CITY_CARBON_XLSX")
    if env_path:
        candidates.append(Path(env_path))
    candidates.extend(
        [
            ROOT / "data" / "地级市绿色金融+碳排放+能源+DID数据（剔除西藏）(1).xlsx",
            Path("D:/CountyCarbonEmissions/地级市绿色金融+碳排放+能源+DID数据（剔除西藏）(1).xlsx"),
        ]
    )
    for path in candidates:
        if path.exists():
            return path
    return candidates[0]


@lru_cache(maxsize=1)
def _load_county_carbon_index() -> dict[tuple[str, int], list[dict[str, Any]]]:
    path = _county_carbon_xlsx_path()
    if not path.exists():
        raise FileNotFoundError(f"县级碳排放数据文件不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        columns = {str(name).strip(): idx for idx, name in enumerate(header) if name is not None}
        required = ["年份", "省份", "地级市", "县级", "CO2排放量_吨"]
        missing = [name for name in required if name not in columns]
        if missing:
            raise ValueError(f"县级碳排放数据缺少字段: {', '.join(missing)}")

        idx: dict[tuple[str, int], list[dict[str, Any]]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            year_raw = row[columns["年份"]]
            province = str(row[columns["省份"]] or "").strip()
            city = str(row[columns["地级市"]] or "").strip()
            county = str(row[columns["县级"]] or "").strip()
            carbon = _safe_float(row[columns["CO2排放量_吨"]])
            if not province or not county or not year_raw:
                continue
            year = int(year_raw)
            item = {
                "year": year,
                "province": province,
                "city": city,
                "county": county,
                "carbonEmission": carbon,
                "carbonEmissionWanTon": round(carbon / 10000, 2),
            }
            idx.setdefault((province, year), []).append(item)

        for rows in idx.values():
            rows.sort(key=lambda item: item["carbonEmission"], reverse=True)
        return idx
    finally:
        workbook.close()


@lru_cache(maxsize=1)
def _load_city_carbon_index() -> dict[tuple[str, int], list[dict[str, Any]]]:
    path = _city_carbon_xlsx_path()
    if not path.exists():
        raise FileNotFoundError(f"地级市碳排放数据文件不存在: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        columns: dict[str, int] = {}
        for idx, name in enumerate(header):
            if name is not None:
                columns.setdefault(str(name).strip(), idx)
        required = ["年份", "省份", "城市代码", "地级市", "CO2排放量"]
        missing = [name for name in required if name not in columns]
        if missing:
            raise ValueError(f"地级市碳排放数据缺少字段: {', '.join(missing)}")

        idx: dict[tuple[str, int], list[dict[str, Any]]] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            year_raw = row[columns["年份"]]
            province = str(row[columns["省份"]] or "").strip()
            city = str(row[columns["地级市"]] or "").strip()
            carbon = _safe_float(row[columns["CO2排放量"]])
            if not province or not city or not year_raw:
                continue
            year = int(year_raw)
            item = {
                "year": year,
                "province": province,
                "city": city,
                "cityCode": row[columns["城市代码"]],
                "carbonEmission": carbon,
                "carbonEmissionWanTon": round(carbon / 10000, 2),
            }
            idx.setdefault((province, year), []).append(item)

        for rows in idx.values():
            rows.sort(key=lambda item: item["carbonEmission"], reverse=True)
        return idx
    finally:
        workbook.close()


@app.get("/api/county-carbon/data")
def get_county_carbon_data(province: str, year: int = 2024, city: Optional[str] = None):
    try:
        idx = _load_county_carbon_index()
        rows = idx.get((province, year), [])
        if city:
            rows = [
                row for row in rows
                if row.get("city") == city
                or str(row.get("city", "")).startswith(city)
                or city.startswith(str(row.get("city", "")))
            ]
        total = sum(_safe_float(row.get("carbonEmission")) for row in rows)
        return {
            "code": 200,
            "msg": "success",
            "data": {
                "province": province,
                "city": city or "",
                "year": year,
                "countyCount": len(rows),
                "carbonEmission": total,
                "carbonEmissionWanTon": round(total / 10000, 2),
                "rows": rows,
            },
        }
    except Exception as e:
        return {"code": 500, "msg": str(e), "data": None}


@app.get("/api/city-carbon/data")
def get_city_carbon_data(province: str, year: int = 2024):
    try:
        idx = _load_city_carbon_index()
        rows = idx.get((province, year), [])
        total = sum(_safe_float(row.get("carbonEmission")) for row in rows)
        return {
            "code": 200,
            "msg": "success",
            "data": {
                "province": province,
                "year": year,
                "cityCount": len(rows),
                "carbonEmission": total,
                "carbonEmissionWanTon": round(total / 10000, 2),
                "rows": rows,
            },
        }
    except Exception as e:
        return {"code": 500, "msg": str(e), "data": None}


@app.get("/api/province/data")
def get_province_data(year: int = 2024):
    rows = get_province_rows(year)
    if rows is None:
        return {"code": 500, "msg": "数据库连接失败", "data": []}
    return {"code": 200, "msg": "success", "data": rows}


@app.get("/api/city/data")
def get_city_data(province: str, year: int = 2024):
    rows = get_city_rows(province, year)
    if rows is None:
        return {"code": 500, "msg": "数据库连接失败", "data": []}
    if not rows:
        return {"code": 200, "msg": f"暂未获取到 {province} 的地级市数据", "data": []}
    return {"code": 200, "msg": "success", "data": rows}


@app.get("/api/macro/cities")
def get_macro_cities(province: Optional[str] = None):
    conn = get_db_connection()
    if not conn:
        return {"code": 500, "msg": "数据库连接失败", "data": []}

    try:
        cursor = conn.cursor()
        if province and province != "全国":
            cursor.execute(
                """
                SELECT DISTINCT province, city
                FROM city_carbon_gdp
                WHERE province=%s
                ORDER BY city ASC
                """,
                (province,),
            )
        else:
            cursor.execute(
                """
                SELECT DISTINCT province, city
                FROM city_carbon_gdp
                ORDER BY province ASC, city ASC
                """
            )
        data = cursor.fetchall()
    except Exception as e:
        print(f"Macro City Error: {e}")
        data = []
    finally:
        conn.close()

    return {"code": 200, "msg": "success", "data": data}


@app.get("/api/macro/data")
<<<<<<< HEAD
def get_macro_data(province: str | None = None):
    rows = get_macro_rows(province)
    if rows is None:
        return {"code": 500, "msg": "数据库连接失败", "data": []}
    return {"code": 200, "msg": "success", "data": rows}
=======
def get_macro_data(province: Optional[str] = None, city: Optional[str] = None):
    conn = get_db_connection()
    if not conn:
        return {"code": 500, "msg": "数据库连接失败", "data": []}

    cursor = conn.cursor()

    carbon_needs_wanton = True
    if city:
        carbon_needs_wanton = False
        if province and province != "全国":
            sql = """
            SELECT year, gdp / 10000 AS gdp, co2_emission / 10000 AS carbonEmission
            FROM city_carbon_gdp
            WHERE province=%s AND city=%s
            ORDER BY year ASC
            """
            cursor.execute(sql, (province, city))
        else:
            sql = """
            SELECT year, gdp / 10000 AS gdp, co2_emission / 10000 AS carbonEmission
            FROM city_carbon_gdp
            WHERE city=%s
            ORDER BY year ASC
            """
            cursor.execute(sql, (city,))
    elif province and province != "全国":
        sql = """
        SELECT year, gdp, carbonEmission
        FROM province_green_finance
        WHERE province=%s
        ORDER BY year ASC
        """
        cursor.execute(sql, (province,))
    else:
        sql = """
        SELECT year, SUM(gdp) as gdp, SUM(carbonEmission) as carbonEmission
        FROM province_green_finance
        GROUP BY year
        ORDER BY year ASC
        """
        cursor.execute(sql)

    data = cursor.fetchall()
    conn.close()

    for row in data:
        if carbon_needs_wanton and row.get("carbonEmission"):
            row["carbonEmission"] = round(row["carbonEmission"] / 10000, 2)
        elif row.get("carbonEmission") is not None:
            row["carbonEmission"] = round(float(row["carbonEmission"]), 2)
        if row.get("gdp") is not None:
            row["gdp"] = round(float(row["gdp"]), 2)

    return {"code": 200, "msg": "success", "data": data}


# ---------------------------------------------------------------------------
# SDM 系数：sdm_coefficients → 前端五维滑块 + rho（与 BizCarbonPrediction.vue 对齐）
# ---------------------------------------------------------------------------
def _load_sdm_coef_pairs() -> tuple[dict[str, float], list[tuple[str, float]]]:
    conn = get_db_connection()
    if not conn:
        return {}, []
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT parameter, value FROM sdm_coefficients ORDER BY id ASC"
        )
        rows = cursor.fetchall()
    finally:
        conn.close()

    raw: dict[str, float] = {}
    ordered: list[tuple[str, float]] = []
    for row in rows:
        name = str(row["parameter"]).strip()
        raw[name] = float(row["value"])
        ordered.append((name, float(row["value"])))
    return raw, ordered


def _build_coefficients_for_frontend(raw: dict[str, float], ordered: list[tuple[str, float]]) -> dict[str, Any]:
    def pick(*names: str) -> float | None:
        for n in names:
            if n in raw:
                return float(raw[n])
        return None

    rho = pick("W_碳排放强度") or (ordered[1][1] if len(ordered) > 1 else 0.4882)
    core = pick("gfi_std") or (ordered[2][1] if len(ordered) > 2 else 0.1446)
    ln_pop = pick("ln_pop") or (ordered[3][1] if len(ordered) > 3 else 1.1873)
    energy = pick("能源强度") or (ordered[4][1] if len(ordered) > 4 else 1.251)
    mediator = pick("人均能源消耗") or (ordered[5][1] if len(ordered) > 5 else -0.3284)
    spatial = pick("W_gfi_std") or (ordered[6][1] if len(ordered) > 6 else -0.0284)

    return {
        "core": core,
        "control": energy,
        "control_ln_pop": ln_pop,
        "policy": float(pick("DID") or pick("did") or 0.0),
        "spatial": spatial,
        "rho": rho,
        "mediator": mediator,
        "raw": raw,
    }
>>>>>>> origin/main


@app.get("/api/dashboard/predict-data")
def get_predict_data():
    payload = get_predict_payload()
    if payload is None:
        return {"code": 500, "msg": "数据库连接失败", "data": None}
    return {"code": 200, "msg": "success", "data": payload}


@app.get("/api/dashboard/macro-stats")
def get_macro_stats():
    records = get_macro_stats_records()
    if records is None:
        return {"code": 500, "msg": "数据库连接失败", "data": None}
    return {"code": 200, "msg": "success", "data": records}


@app.post("/api/ai/chat")
def ai_chat(payload: AiChatRequest):
    try:
        result = request_deepseek_completion(
            build_chat_messages(payload),
            temperature=0.3,
        )
        return {
            "code": 200,
            "msg": "success",
            "data": {
                "content": result["content"],
                "model": result["model"],
                "kind": "chat",
            },
        }
    except AiServiceError as exc:
        return {"code": 500, "msg": str(exc), "data": None}
    except Exception as exc:  # pragma: no cover
        return {"code": 500, "msg": f"AI 聊天调用失败: {exc}", "data": None}


@app.post("/api/ai/summary")
def ai_summary(payload: AiSummaryRequest):
    try:
        result = request_deepseek_completion(
            build_summary_messages(payload),
            temperature=0.4,
        )
        return {
            "code": 200,
            "msg": "success",
            "data": {
                "content": result["content"],
                "model": result["model"],
                "kind": "summary",
            },
        }
    except AiServiceError as exc:
        return {"code": 500, "msg": str(exc), "data": None}
    except Exception as exc:  # pragma: no cover
        return {"code": 500, "msg": f"AI 总结调用失败: {exc}", "data": None}


@app.post("/api/ai/chat/stream")
def ai_chat_stream(payload: AiChatRequest):
    return StreamingResponse(
        stream_chat_agent_events(payload),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/ai/summary/stream")
def ai_summary_stream(payload: AiSummaryRequest):
    return StreamingResponse(
        stream_summary_agent_events(payload),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )
